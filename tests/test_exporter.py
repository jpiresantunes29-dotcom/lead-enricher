"""
Testes do exportador XLSX/CSV.

Regressão: as colunas de decisor liam atributos inexistentes no Lead (saíam
sempre vazias) e employee_count era exportado como o dict cru do JSON.
"""
import csv
import io
from datetime import datetime, UTC

from openpyxl import load_workbook

from models.database import Lead, DecisionMaker
from services.exporter import build_csv, build_excel, HEADERS, _format_employee_count


def _lead(**over):
    lead = Lead(
        id=1,
        raw_input_domain="acme.com.br",
        domain="acme.com.br",
        company_name="Acme Tecnologia",
        website="https://acme.com.br",
        sector="Tecnologia",
        location="São Paulo, SP",
        status="enriched",
        created_at=datetime(2026, 7, 5, 14, 30, tzinfo=UTC),
    )
    for k, v in over.items():
        setattr(lead, k, v)
    return lead


def _csv_rows(leads):
    # parser real: campos como "São Paulo, SP" vêm entre aspas
    text = build_csv(leads).decode("utf-8-sig")
    return list(csv.reader(io.StringIO(text)))


# ── employee_count ────────────────────────────────────────────────────────────
def test_format_employee_count_variants():
    assert _format_employee_count({"exact": 320}) == "320"
    assert _format_employee_count({"min": 51, "max": 200}) == "51–200"
    assert _format_employee_count({"min": 5001}) == "5001+"
    assert _format_employee_count({"band": "11-50"}) == "11-50"
    assert _format_employee_count({"raw": "cerca de 40"}) == "cerca de 40"
    assert _format_employee_count(None) == ""


def test_csv_employee_count_is_readable_not_dict():
    lead = _lead(employee_count={"raw": "51-200 funcionários", "min": 51, "max": 200,
                                 "band": "51-200", "source": "linkedin_direct"})
    rows = _csv_rows([lead])
    idx = [f for _, f in HEADERS].index("employee_count")
    assert rows[1][idx] == "51–200"
    assert "{" not in rows[1][idx]


# ── colunas de decisor ────────────────────────────────────────────────────────
def test_csv_includes_first_decision_maker():
    lead = _lead()
    lead.decision_makers = [
        DecisionMaker(
            name="Marina Alves",
            title_found="Diretora de TI",
            linkedin_url="https://linkedin.com/in/marina",
        )
    ]
    rows = _csv_rows([lead])
    fields = [f for _, f in HEADERS]
    assert rows[1][fields.index("decision_maker_name")] == "Marina Alves"
    assert rows[1][fields.index("decision_maker_title")] == "Diretora de TI"
    assert rows[1][fields.index("decision_maker_linkedin")] == "https://linkedin.com/in/marina"


def test_decision_maker_title_falls_back_to_searched():
    lead = _lead()
    lead.decision_makers = [DecisionMaker(name="Rafael Nunes", title_searched="CFO")]
    rows = _csv_rows([lead])
    fields = [f for _, f in HEADERS]
    assert rows[1][fields.index("decision_maker_title")] == "CFO"


def test_lead_without_decision_makers_exports_empty_columns():
    lead = _lead()
    lead.decision_makers = []
    rows = _csv_rows([lead])
    fields = [f for _, f in HEADERS]
    assert rows[1][fields.index("decision_maker_name")] == ""


# ── estrutura dos arquivos ────────────────────────────────────────────────────
def test_csv_header_matches_headers_definition():
    rows = _csv_rows([_lead()])
    assert rows[0] == [label for label, _ in HEADERS]


def test_excel_is_valid_and_has_data(monkeypatch):
    lead = _lead(employee_count={"exact": 320})
    lead.decision_makers = [DecisionMaker(name="Marina Alves", title_found="Diretora de TI")]
    wb = load_workbook(io.BytesIO(build_excel([lead])))
    ws = wb.active
    values = [c.value for c in ws[2]]
    assert "Acme Tecnologia" in values
    assert "320" in values
    assert "Marina Alves" in values
    assert ws.freeze_panes == "A2"
