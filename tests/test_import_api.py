"""
Testes de integração da importação de planilha e da view de planilha.

Reusa o banco SQLite em memória e o mock de autenticação de tests/test_api.py.
"""
import io
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from main import app
from models.database import ImportBatch, Lead, Profile
from tests.test_api import (  # noqa: F401  (clean_db é autouse)
    FAKE_USER, _Session, clean_db,
)

# Rate limiter desligado: a suíte estoura o limite por minuto
from routers.imports import limiter as _import_limiter  # noqa: E402
_import_limiter.enabled = False
from routers.enrichment import limiter as _enrich_limiter  # noqa: E402
_enrich_limiter.enabled = False


@pytest.fixture
def client():
    from middleware.auth import get_current_user
    app.dependency_overrides[get_current_user] = lambda: FAKE_USER
    with TestClient(app) as c:
        yield c
    app.dependency_overrides.pop(get_current_user, None)


def _xlsx(rows, sheets=None):
    wb = Workbook()
    wb.remove(wb.active)
    for name, data in (sheets or {"Base": rows}).items():
        ws = wb.create_sheet(title=name)
        for row in data:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


BASE_ROWS = [
    ["🏢 Empresa", "🔗 LinkedIn", "🏆 TIER", "✅ Status", "📩 Email"],
    ["Nubank", "https://linkedin.com/company/nubank", "TIER 1", "Qualificado", ""],
    ["TOTVS", "https://linkedin.com/company/totvs", "TIER 2", "Qualificado", "contato@totvs.com.br"],
]


def _upload(client, content=None, filename="base.xlsx", sheet=None):
    content = content if content is not None else _xlsx(BASE_ROWS)
    return client.post(
        "/api/import/preview",
        files={"file": (filename, io.BytesIO(content), "application/octet-stream")},
        data={"sheet": sheet} if sheet else None,
    )


def _import(client, content=None, **commit):
    preview = _upload(client, content).json()
    resp = client.post(f"/api/import/{preview['batch_id']}/commit", json=commit or {})
    return preview, resp.json()


# ── preview ───────────────────────────────────────────────────────────────────
def test_preview_devolve_colunas_na_ordem_e_nao_grava_leads(client):
    data = _upload(client).json()

    assert data["importable"] == 2
    assert [c["label"] for c in data["columns"]] == [
        "🏢 Empresa", "🔗 LinkedIn", "🏆 TIER", "✅ Status", "📩 Email"]
    db = _Session()
    assert db.query(Lead).count() == 0
    assert db.query(ImportBatch).filter(ImportBatch.status == "draft").count() == 1
    db.close()


def test_preview_lista_abas_e_escolhe_a_de_leads(client):
    content = _xlsx(None, sheets={
        "Controle": [["Data", "Total"], ["2026-01-01", 10]],
        "Leads": BASE_ROWS,
    })
    data = _upload(client, content).json()
    assert data["sheet"] == "Leads"
    assert {s["name"] for s in data["sheets"]} == {"Controle", "Leads"}


def test_preview_aceita_aba_escolhida(client):
    content = _xlsx(None, sheets={"Leads": BASE_ROWS, "Outra": [["Empresa"], ["Beta"]]})
    data = _upload(client, content, sheet="Outra").json()
    assert data["sheet"] == "Outra"
    assert data["total_rows"] == 1


def test_preview_marca_empresa_que_ja_esta_no_historico(client):
    db = _Session()
    db.add(Lead(user_id=FAKE_USER["sub"], raw_input_domain="Nubank",
                company_name="Nubank", status="enriched"))
    db.commit()
    db.close()

    data = _upload(client).json()
    assert data["counts"]["duplicate_db"] == 1
    assert data["importable"] == 1


def test_preview_substitui_rascunho_anterior(client):
    _upload(client)
    _upload(client)
    db = _Session()
    assert db.query(ImportBatch).filter(ImportBatch.status == "draft").count() == 1
    db.close()


def test_preview_rejeita_formato_invalido(client):
    assert _upload(client, b"%PDF-1.4", filename="lista.pdf").status_code == 422


# ── commit ────────────────────────────────────────────────────────────────────
def test_commit_guarda_todas_as_celulas_originais(client):
    _, result = _import(client)
    assert result["created"] == 2

    db = _Session()
    lead = db.query(Lead).filter(Lead.company_name == "Nubank").one()
    assert lead.cells["🏆 TIER"] == "TIER 1"
    assert lead.cells["✅ Status"] == "Qualificado"
    assert lead.cells["🔗 LinkedIn"] == "https://linkedin.com/company/nubank"
    assert lead.status == "imported"
    assert lead.sheet_row == 2
    assert lead.import_batch_id == result["batch_id"]
    db.close()


def test_commit_nao_consome_cota(client):
    db = _Session()
    db.add(Profile(id=FAKE_USER["sub"], plan="free", searches_used=0, searches_limit=5))
    db.commit()
    db.close()

    _import(client)

    db = _Session()
    assert db.get(Profile, FAKE_USER["sub"]).searches_used == 0
    db.close()


def test_commit_limpa_o_rascunho(client):
    _, result = _import(client)
    db = _Session()
    batch = db.get(ImportBatch, result["batch_id"])
    assert batch.status == "committed"
    assert batch.draft_rows is None
    db.close()


def test_commit_duas_vezes_nao_duplica(client):
    preview, _ = _import(client)
    again = client.post(f"/api/import/{preview['batch_id']}/commit", json={})
    assert again.status_code == 409
    db = _Session()
    assert db.query(Lead).count() == 2
    db.close()


def test_commit_pode_pular_duplicadas_do_arquivo(client):
    content = _xlsx([["Empresa"], ["Acme"], ["ACME"]])
    _, com_duplicadas = _import(client, content)
    assert com_duplicadas["created"] == 2      # padrão: preserva as duas linhas

    _, sem_duplicadas = _import(client, content, skip_duplicates=True, skip_existing=False)
    assert sem_duplicadas["created"] == 1


def test_batches_lista_e_conta(client):
    _import(client)
    batches = client.get("/api/import/batches").json()
    assert len(batches) == 1
    assert batches[0]["row_count"] == 2
    assert batches[0]["enriched_count"] == 0


def test_excluir_lote_remove_os_leads(client):
    _, result = _import(client)
    assert client.delete(f"/api/import/batches/{result['batch_id']}").status_code == 204
    db = _Session()
    assert db.query(Lead).count() == 0
    assert db.query(ImportBatch).count() == 0
    db.close()


# ── planilha ──────────────────────────────────────────────────────────────────
def test_sheet_devolve_colunas_do_arquivo_e_do_sistema(client):
    _import(client)
    data = client.get("/api/sheet").json()

    assert [c["label"] for c in data["columns"]][:3] == ["🏢 Empresa", "🔗 LinkedIn", "🏆 TIER"]
    assert "Domínio" in [c["label"] for c in data["system_columns"]]
    assert data["total"] == 2
    assert data["enrichable"] == 2
    row = data["rows"][0]
    assert row["values"]["🏆 TIER"] == "TIER 1"
    assert row["values"]["Situação"] == "Importado"


def test_sheet_busca_em_qualquer_coluna(client):
    _import(client)
    assert client.get("/api/sheet", params={"q": "TIER 2"}).json()["total"] == 1
    assert client.get("/api/sheet", params={"q": "nubank"}).json()["total"] == 1
    assert client.get("/api/sheet", params={"q": "inexistente"}).json()["total"] == 0


def test_sheet_ordena_por_coluna_do_arquivo(client):
    content = _xlsx([["Empresa", "Funcionários"], ["A", 39], ["B", "1000 a 5000"], ["C", 200]])
    _import(client, content)
    rows = client.get("/api/sheet", params={"sort": "Funcionários", "dir": "desc"}).json()["rows"]
    assert [r["values"]["Empresa"] for r in rows] == ["B", "C", "A"]


def test_sheet_ordena_mantendo_vazios_no_fim(client):
    content = _xlsx([["Empresa", "Nota"], ["A", "10"], ["B", None], ["C", "30"]])
    _import(client, content)
    for direcao in ("asc", "desc"):
        rows = client.get("/api/sheet", params={"sort": "Nota", "dir": direcao}).json()["rows"]
        assert rows[-1]["values"]["Empresa"] == "B"


def test_sheet_pagina(client):
    content = _xlsx([["Empresa"]] + [[f"Empresa {i}"] for i in range(10)])
    _import(client, content)
    page2 = client.get("/api/sheet", params={"per_page": 4, "page": 2}).json()
    assert page2["total"] == 10
    assert len(page2["rows"]) == 4


def test_sheet_filtra_pendentes_e_enriquecidas(client):
    _import(client)
    db = _Session()
    lead = db.query(Lead).first()
    lead.status = "enriched"
    db.commit()
    db.close()

    assert client.get("/api/sheet", params={"only": "pending"}).json()["total"] == 1
    assert client.get("/api/sheet", params={"only": "enriched"}).json()["total"] == 1


# ── edição de célula ──────────────────────────────────────────────────────────
def _first_lead_id(client):
    return client.get("/api/sheet").json()["rows"][0]["id"]


def test_edita_celula_da_planilha(client):
    _import(client)
    lead_id = _first_lead_id(client)
    resp = client.patch("/api/sheet/cell",
                        json={"lead_id": lead_id, "column": "🏆 TIER", "value": "TIER 3"})
    assert resp.status_code == 200
    assert resp.json()["row"]["values"]["🏆 TIER"] == "TIER 3"

    db = _Session()
    assert db.get(Lead, lead_id).cells["🏆 TIER"] == "TIER 3"
    db.close()


def test_edita_coluna_do_sistema_normalizando(client):
    _import(client)
    lead_id = _first_lead_id(client)
    resp = client.patch("/api/sheet/cell", json={
        "lead_id": lead_id, "column": "Domínio", "value": "https://www.nubank.com.br/pj"})
    values = resp.json()["row"]["values"]
    assert values["Domínio"] == "nubank.com.br"
    assert values["Site"] == "https://nubank.com.br"


def test_coluna_calculada_nao_e_editavel(client):
    _import(client)
    resp = client.patch("/api/sheet/cell",
                        json={"lead_id": _first_lead_id(client), "column": "Prioridade", "value": "Alta"})
    assert resp.status_code == 422


def test_celula_vazia_some_do_registro(client):
    _import(client)
    lead_id = _first_lead_id(client)
    client.patch("/api/sheet/cell", json={"lead_id": lead_id, "column": "🏆 TIER", "value": ""})
    db = _Session()
    assert "🏆 TIER" not in (db.get(Lead, lead_id).cells or {})
    db.close()


def test_nao_edita_linha_de_outro_usuario(client):
    db = _Session()
    lead = Lead(user_id="outro", raw_input_domain="x.com", status="imported")
    db.add(lead)
    db.commit()
    lead_id = lead.id
    db.close()

    resp = client.patch("/api/sheet/cell",
                        json={"lead_id": lead_id, "column": "Domínio", "value": "y.com"})
    assert resp.status_code == 404


# ── linhas ────────────────────────────────────────────────────────────────────
def test_cria_e_exclui_linha(client):
    _, result = _import(client)
    resp = client.post("/api/sheet/row", json={
        "batch_id": result["batch_id"],
        "values": {"🏢 Empresa": "Nova Empresa", "Domínio": "nova.com.br"},
    })
    assert resp.status_code == 200
    row = resp.json()["row"]
    assert row["values"]["🏢 Empresa"] == "Nova Empresa"
    assert row["values"]["Domínio"] == "nova.com.br"
    assert client.get("/api/sheet").json()["total"] == 3

    assert client.post("/api/sheet/rows/delete", json={"ids": [row["id"]]}).status_code == 200
    assert client.get("/api/sheet").json()["total"] == 2


# ── exportação ────────────────────────────────────────────────────────────────
def test_export_traz_colunas_originais_e_do_sistema(client):
    _import(client)
    resp = client.get("/api/sheet/export")
    assert resp.status_code == 200

    ws = load_workbook(io.BytesIO(resp.content)).active
    header = [c.value for c in ws[1]]
    assert header[:5] == ["🏢 Empresa", "🔗 LinkedIn", "🏆 TIER", "✅ Status", "📩 Email"]
    assert "Domínio" in header and "Prioridade" in header
    assert ws.max_row == 3          # cabeçalho + 2 linhas
    assert ws.cell(row=2, column=3).value == "TIER 1"


def test_export_respeita_a_busca(client):
    _import(client)
    resp = client.get("/api/sheet/export", params={"q": "TIER 2"})
    ws = load_workbook(io.BytesIO(resp.content)).active
    assert ws.max_row == 2


# ── enriquecimento a partir da planilha ──────────────────────────────────────
MOCK_RESULT = {
    "raw_input_domain": "totvs.com.br", "domain": "totvs.com.br",
    "website": "https://totvs.com.br", "company_name": "TOTVS S.A.",
    "description": "ERP", "location": "São Paulo, Brasil", "sector": "Tecnologia",
    "corporate_email": None, "phone": None,
    "linkedin_url": "https://linkedin.com/company/totvs", "linkedin_confidence": "verified",
    "mx_provider": "Google", "mx_provider_confidence": "high", "mx_records": [],
    "dns_report": None, "hosting_provider": "AWS", "employee_count": None,
    "employee_count_linkedin": None, "status": "enriched",
}


def test_enriquecer_nao_toca_nas_celulas_da_planilha(client):
    """O arquivo do usuário continua intacto; a coleta só preenche colunas do sistema."""
    _import(client)
    db = _Session()
    lead = db.query(Lead).filter(Lead.company_name == "TOTVS").one()
    lead_id, cells_antes = lead.id, dict(lead.cells)
    db.close()

    with patch("services.enrichment_service.enrich_company", return_value=MOCK_RESULT):
        resp = client.post(f"/api/leads/{lead_id}/enrich")
    assert resp.status_code == 200

    db = _Session()
    lead = db.get(Lead, lead_id)
    assert lead.cells == cells_antes            # planilha do usuário, intacta
    assert lead.company_name == "TOTVS S.A."    # coluna do sistema, atualizada
    assert lead.status == "enriched"
    db.close()

    rows = client.get("/api/sheet").json()["rows"]
    row = next(r for r in rows if r["id"] == lead_id)
    assert row["values"]["🏢 Empresa"] == "TOTVS"                # coluna do arquivo
    assert row["values"]["Empresa (coletada)"] == "TOTVS S.A."   # coluna do sistema


def test_enriquecer_descobre_dominio_pelo_nome(client):
    """Linha só com nome e LinkedIn: o domínio é descoberto antes da coleta."""
    _import(client)
    db = _Session()
    lead_id = db.query(Lead).filter(Lead.company_name == "Nubank").one().id
    db.close()

    with patch("services.enrichment_service.find_domain",
               return_value={"domain": "nubank.com.br", "confidence": "high", "evidence": "x"}) as finder, \
         patch("services.enrichment_service.enrich_company", return_value=MOCK_RESULT) as coleta:
        resp = client.post(f"/api/leads/{lead_id}/enrich")

    assert resp.status_code == 200
    # Descobriu pelo nome e coletou o domínio descoberto
    assert finder.call_args.args[0] == "Nubank"
    assert coleta.call_args.args[0] == "nubank.com.br"


def test_sem_dominio_descoberto_a_linha_falha_com_recado(client):
    _import(client)
    db = _Session()
    lead_id = db.query(Lead).filter(Lead.company_name == "Nubank").one().id
    db.close()

    with patch("services.enrichment_service.find_domain",
               return_value={"domain": None, "confidence": "none", "evidence": "nada"}):
        resp = client.post(f"/api/leads/{lead_id}/enrich")

    assert resp.status_code == 422
    assert "Domínio" in resp.json()["detail"]
    db = _Session()
    assert db.get(Lead, lead_id).status == "failed"
    db.close()


def test_enriquecer_consome_uma_busca_por_linha(client):
    db = _Session()
    db.add(Profile(id=FAKE_USER["sub"], plan="free", searches_used=0, searches_limit=5))
    db.commit()
    db.close()
    _import(client)
    db = _Session()
    lead_id = db.query(Lead).filter(Lead.company_name == "TOTVS").one().id
    db.close()

    with patch("services.enrichment_service.enrich_company", return_value=MOCK_RESULT):
        client.post(f"/api/leads/{lead_id}/enrich")

    db = _Session()
    assert db.get(Profile, FAKE_USER["sub"]).searches_used == 1
    db.close()


def test_fila_para_quando_a_cota_acaba(client):
    db = _Session()
    db.add(Profile(id=FAKE_USER["sub"], plan="free", searches_used=5, searches_limit=5))
    db.commit()
    db.close()
    _import(client)
    db = _Session()
    lead_id = db.query(Lead).filter(Lead.company_name == "TOTVS").one().id
    db.close()

    assert client.post(f"/api/leads/{lead_id}/enrich").status_code == 402


# ── modelo de planilha ────────────────────────────────────────────────────────
@pytest.mark.parametrize("fmt", ["xlsx", "csv"])
def test_download_do_modelo(client, fmt):
    resp = client.get(f"/api/import/template?format={fmt}")
    assert resp.status_code == 200
    assert "modelo_importacao" in resp.headers["content-disposition"]
