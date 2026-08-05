"""
Importação de planilhas de empresas (XLSX/CSV) com fidelidade total.

Princípio: **nenhuma célula é descartada**. Toda coluna do arquivo vira uma
coluna da planilha dentro do sistema, com o rótulo e a ordem originais; as
colunas que o sistema reconhece (empresa, domínio, LinkedIn, e-mail…) também
alimentam os campos do Lead, que é o que o enriquecimento usa.

Os valores são preservados por tipo: data vira ISO 8601, número continua
número, texto (inclusive multi-linha, com acento e emoji) fica byte a byte
como está no arquivo. O que o sistema deduz — domínio a partir do e-mail, por
exemplo — vai para os campos do Lead, nunca por cima da célula original.
"""
import csv
import io
import re
import unicodedata
from datetime import date, datetime, time
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from ._utils import LINKEDIN_COMPANY_RE, normalize_domain

# Limites — protegem a função serverless (60 s na Vercel) e a memória.
MAX_FILE_BYTES = 15 * 1024 * 1024   # 15 MB
MAX_ROWS = 5000                     # linhas de dados por importação
HEADER_SCAN_ROWS = 10               # até onde procuramos a linha de cabeçalho

XLSX_EXTENSIONS = (".xlsx", ".xlsm")
CSV_EXTENSIONS = (".csv", ".tsv", ".txt")

# Colunas que o sistema entende. A chave é o campo do Lead; os aliases são
# comparados com o cabeçalho normalizado (sem acento, sem emoji, minúsculo).
# Match exato tem prioridade sobre prefixo/sufixo — sem isso "Servidor de
# Email" roubaria a coluna de e-mail de contato.
FIELD_ALIASES: Dict[str, Tuple[str, ...]] = {
    "domain": ("dominio", "domain", "site", "website", "url", "pagina web", "web site"),
    "company_name": (
        "empresa", "nome", "nome da empresa", "razao social", "razaosocial",
        "company", "company name", "cliente", "conta", "account",
    ),
    "linkedin_url": ("linkedin", "linkedin url", "linkedin da empresa", "perfil linkedin"),
    "corporate_email": ("email", "e mail", "email corporativo", "email de contato", "mail"),
    "phone": ("telefone", "fone", "celular", "whatsapp", "phone", "tel", "contato"),
    "sector": ("setor", "segmento", "industria", "ramo", "sector", "industry"),
    "location": ("localizacao", "local", "cidade", "endereco", "location", "city"),
    "employee_count": (
        "funcionarios", "n funcionarios", "no funcionarios", "numero de funcionarios",
        "colaboradores", "porte", "employees", "headcount",
    ),
    "mx_provider": (
        "servidor de email", "provedor de email", "email atual", "servidor",
        "mx", "provedor", "email provider",
    ),
    "description": ("descricao", "sobre", "description", "about"),
}

# Campos de texto simples copiados direto para o Lead
TEXT_FIELDS = (
    "company_name", "sector", "location", "description",
    "corporate_email", "phone", "linkedin_url", "mx_provider",
)

# Limites das colunas VARCHAR correspondentes em models.database.Lead
_MAX_LEN = {
    "company_name": 255, "sector": 255, "location": 255, "corporate_email": 255,
    "phone": 100, "linkedin_url": 500, "mx_provider": 100, "description": 5000,
}

_DOMAIN_RE = re.compile(r"^[a-z0-9](?:[a-z0-9-]*[a-z0-9])?(?:\.[a-z0-9-]+)+$")
_EMAIL_RE = re.compile(r"[\w.+-]+@([\w-]+\.[\w.-]+)")

TEMPLATE_HEADERS = [
    ("Domínio", "domain"),
    ("Empresa", "company_name"),
    ("Setor", "sector"),
    ("Localização", "location"),
    ("Email Corporativo", "corporate_email"),
    ("Telefone", "phone"),
    ("LinkedIn", "linkedin_url"),
    ("Funcionários", "employee_count"),
    ("Descrição", "description"),
]

TEMPLATE_SAMPLE = [
    ["nubank.com.br", "Nubank", "Serviços financeiros", "São Paulo, SP",
     "contato@nubank.com.br", "+55 11 4000-0000",
     "https://linkedin.com/company/nubank", "5001-10000", "Banco digital"],
    ["totvs.com.br", "TOTVS", "Tecnologia", "São Paulo, SP", "", "", "", "1001-5000", ""],
]


class SpreadsheetError(Exception):
    """Erro de parsing tratável — o router converte em HTTP 422."""


# ── normalização de rótulos e valores ─────────────────────────────────────────

def _strip_accents(value: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(c)
    )


def _normalize_header(value: Any) -> str:
    """'📩 Email Decisor 2' → 'email decisor 2' (emoji e acento fora)."""
    text = _strip_accents(str(value or "")).lower().strip()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _cell_value(value: Any) -> Any:
    """
    Converte a célula para algo que sobrevive ao JSON sem perder informação:
    data/hora vira ISO 8601, número continua número, texto fica intacto.
    """
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime):
        # Meia-noite exata: a planilha guarda só a data
        if value.time() == time(0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    return text if text.strip() else None


def _cell_text(value: Any) -> str:
    """Versão texto da célula, para casar cabeçalho e alimentar campos do Lead."""
    parsed = _cell_value(value)
    if parsed is None:
        return ""
    return str(parsed).strip()


def _kind_of(value: Any) -> str:
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


# ── leitura bruta ─────────────────────────────────────────────────────────────

def _read_xlsx_sheets(content: bytes) -> Dict[str, List[List[Any]]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise SpreadsheetError(f"Não consegui ler o arquivo Excel: {e}")
    try:
        sheets: Dict[str, List[List[Any]]] = {}
        limit = MAX_ROWS + HEADER_SCAN_ROWS + 1
        for name in wb.sheetnames:
            ws = wb[name]
            sheets[name] = [list(row) for row in ws.iter_rows(values_only=True, max_row=limit)]
        return sheets
    finally:
        wb.close()


def _read_csv_sheet(content: bytes) -> Dict[str, List[List[Any]]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise SpreadsheetError("Não consegui identificar a codificação do arquivo.")
    sample = text[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        delimiter = max(",;\t|", key=sample.count) if sample else ","
        if sample.count(delimiter) == 0:
            delimiter = ","
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    limit = MAX_ROWS + HEADER_SCAN_ROWS + 1
    return {"Planilha": [row for _, row in zip(range(limit), reader)]}


def _read_sheets(filename: str, content: bytes) -> Dict[str, List[List[Any]]]:
    if len(content) > MAX_FILE_BYTES:
        raise SpreadsheetError(
            f"Arquivo muito grande (máx. {MAX_FILE_BYTES // (1024 * 1024)} MB)."
        )
    if not content:
        raise SpreadsheetError("Arquivo vazio.")
    name = (filename or "").lower()
    if name.endswith(XLSX_EXTENSIONS):
        return _read_xlsx_sheets(content)
    if name.endswith(CSV_EXTENSIONS):
        return _read_csv_sheet(content)
    if name.endswith(".xls"):
        raise SpreadsheetError(
            "Formato .xls (Excel 97-2003) não suportado. Salve como .xlsx ou .csv."
        )
    raise SpreadsheetError("Formato não suportado. Envie um arquivo .xlsx ou .csv.")


# ── cabeçalho e mapeamento ────────────────────────────────────────────────────

def _map_columns(header_row: List[Any]) -> Dict[int, str]:
    """
    Casa colunas do cabeçalho com campos do Lead: primeiro os nomes exatos,
    depois prefixo/sufixo. Cada campo é usado por uma coluna só — a primeira
    que casar de forma mais específica fica com ele.
    """
    labels = [_cell_text(c) for c in header_row]
    norms = [_normalize_header(label) for label in labels]
    mapping: Dict[int, str] = {}
    taken: set[str] = set()

    for field, aliases in FIELD_ALIASES.items():
        for idx, norm in enumerate(norms):
            if not norm or idx in mapping or field in taken:
                continue
            if norm in aliases:
                mapping[idx] = field
                taken.add(field)

    for field, aliases in FIELD_ALIASES.items():
        if field in taken:
            continue
        for idx, norm in enumerate(norms):
            if not norm or idx in mapping:
                continue
            if any(norm.startswith(a + " ") or norm.endswith(" " + a) for a in aliases):
                mapping[idx] = field
                taken.add(field)
                break
    return mapping


def _find_header_row(rows: List[List[Any]]) -> int:
    """
    Cabeçalho é a linha que reconhece mais campos nas primeiras linhas — o que
    pula título, logo e linhas em branco que planilhas reais costumam ter.
    """
    best_idx, best_score = -1, 0
    for idx, row in enumerate(rows[:HEADER_SCAN_ROWS]):
        if not any(_cell_text(c) for c in row):
            continue
        score = len(_map_columns(row))
        if score > best_score:
            best_idx, best_score = idx, score
    if best_idx >= 0:
        return best_idx
    for idx, row in enumerate(rows):
        if any(_cell_text(c) for c in row):
            return idx
    raise SpreadsheetError("A planilha não tem nenhuma linha preenchida.")


def _unique_labels(header_row: List[Any]) -> List[str]:
    """Rótulos originais, com sufixo em duplicatas e nome para coluna sem título."""
    labels: List[str] = []
    seen: Dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        label = _cell_text(cell) or f"Coluna {idx + 1}"
        if label in seen:
            seen[label] += 1
            label = f"{label} ({seen[label]})"
        else:
            seen[label] = 1
        labels.append(label)
    return labels


# ── extração de identidade ────────────────────────────────────────────────────

def _slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", _strip_accents(str(value or "")).lower())


def domain_matches_company(domain: str, company_name: str) -> bool:
    """
    O domínio do e-mail é da empresa da linha, ou é de outra?

    Planilhas de prospecção acumulam e-mails colados na linha errada. Herdar
    um domínio desses enriqueceria a empresa errada e sujaria a ficha, então
    só aceitamos quando nome e domínio se parecem — o resto vai para a
    descoberta de domínio, que busca pelo nome.
    """
    host = _slug(normalize_domain(domain).split(".")[0])
    name = _slug(company_name)
    if not host or not name:
        return False
    if host in name or name in host:
        return True
    if len(name) >= 5 and name[:5] in host:
        return True
    from difflib import SequenceMatcher
    return SequenceMatcher(None, name, host).ratio() >= 0.55


def _extract_domain(values: Dict[str, str]) -> Tuple[str, str]:
    """
    Devolve (domínio, origem). Da coluna de site vale sempre; do e-mail só
    quando o domínio combina com o nome da empresa.
    """
    candidate = values.get("domain", "")
    if candidate:
        domain = normalize_domain(candidate)
        if _DOMAIN_RE.match(domain):
            return domain, "site"
    match = _EMAIL_RE.search(values.get("corporate_email", ""))
    if match:
        domain = normalize_domain(match.group(1))
        if _DOMAIN_RE.match(domain):
            company = values.get("company_name", "")
            if not company or domain_matches_company(domain, company):
                return domain, "email"
    return "", ""


def _clean_linkedin(value: str) -> str:
    """Mantém só URL de empresa (o /in/ das colunas de decisor não é da empresa)."""
    match = LINKEDIN_COMPANY_RE.search(value or "")
    return match.group(0).rstrip("/") if match else ""


def parse_employee_count(value: Any) -> Optional[dict]:
    """
    '250', '51-200', '1.001+', '200 a 500' → o mesmo formato consolidado que o
    enricher grava ({raw, min, max, band, exact, source}).
    """
    raw = str(value).strip() if value not in (None, "") else ""
    if not raw:
        return None
    digits = [int(n.replace(".", "").replace(",", "")) for n in re.findall(r"\d[\d.,]*", raw)]
    result: Dict[str, Any] = {
        "raw": raw, "min": None, "max": None, "band": None,
        "exact": None, "source": "import",
    }
    if len(digits) >= 2:
        result["min"], result["max"] = digits[0], digits[1]
        result["band"] = f"{digits[0]}-{digits[1]}"
    elif len(digits) == 1:
        if "+" in raw or "mais" in raw.lower():
            result["min"] = digits[0]
            result["band"] = f"{digits[0]}+"
        else:
            result["exact"] = digits[0]
    else:
        result["band"] = raw
    return result


def looks_like_domain(value: str) -> bool:
    """'acme.com.br' sim; 'Acme Tecnologia' não."""
    return bool(_DOMAIN_RE.match(normalize_domain(value or "")))


def identity_keys(lead: Dict[str, Any]) -> set:
    """
    Todas as identidades da linha: domínio, LinkedIn e nome.

    São comparadas por interseção porque a mesma empresa aparece de formas
    diferentes em arquivos diferentes — numa planilha só o LinkedIn, noutra só
    o site. Bastar uma bater já é duplicata.
    """
    keys = set()
    if lead.get("domain"):
        keys.add("d:" + lead["domain"])
    if lead.get("linkedin_url"):
        keys.add("l:" + lead["linkedin_url"].lower().rstrip("/"))
    if lead.get("company_name"):
        keys.add("n:" + _normalize_header(lead["company_name"]))
    return keys


def identity_key(lead: Dict[str, Any]) -> str:
    """Identidade preferida da linha (domínio > LinkedIn > nome), ou vazio."""
    for prefix in ("d:", "l:", "n:"):
        for key in identity_keys(lead):
            if key.startswith(prefix):
                return key
    return ""


# ── leitura de uma aba ────────────────────────────────────────────────────────

def _sheet_summary(name: str, rows: List[List[Any]]) -> Dict[str, Any]:
    """Resumo usado para escolher a aba certa quando o arquivo tem várias."""
    if not any(any(_cell_text(c) for c in row) for row in rows):
        return {"name": name, "rows": 0, "columns": 0, "mapped_fields": [], "score": -1}
    header_idx = _find_header_row(rows)
    mapping = _map_columns(rows[header_idx])
    data_rows = [r for r in rows[header_idx + 1:] if any(_cell_text(c) for c in r)]
    fields = sorted(set(mapping.values()))
    identity = {"domain", "company_name", "linkedin_url"} & set(fields)
    return {
        "name": name,
        "rows": len(data_rows),
        "columns": len(_unique_labels(rows[header_idx])),
        "mapped_fields": fields,
        # Aba de leads = tem identidade de empresa e volume; a de métricas não
        "score": (len(identity) * 1000 + len(fields) * 10) if identity else 0,
    }


def list_sheets(filename: str, content: bytes) -> List[Dict[str, Any]]:
    sheets = _read_sheets(filename, content)
    return [_sheet_summary(name, rows) for name, rows in sheets.items()]


def parse_spreadsheet(
    filename: str,
    content: bytes,
    sheet_name: Optional[str] = None,
) -> Dict[str, Any]:
    """
    Lê uma aba inteira preservando todas as células e devolve:
      {sheet, sheets, columns, rows, total_rows, truncated, counts}

    columns: [{label, field, kind}] na ordem original do arquivo.
    rows:    [{row_number, status, reason, cells, lead}]
             cells = {rótulo original: valor} — a planilha do usuário, intacta.
             lead  = campos que alimentam o Lead (domínio, empresa, LinkedIn…).
    """
    sheets = _read_sheets(filename, content)
    summaries = [_sheet_summary(name, rows) for name, rows in sheets.items()]

    if sheet_name and sheet_name not in sheets:
        raise SpreadsheetError(f'A aba "{sheet_name}" não existe neste arquivo.')
    if not sheet_name:
        best = max(summaries, key=lambda s: (s["score"], s["rows"]))
        sheet_name = best["name"]

    raw_rows = sheets[sheet_name]
    if not any(any(_cell_text(c) for c in row) for row in raw_rows):
        raise SpreadsheetError(f'A aba "{sheet_name}" está vazia.')

    header_idx = _find_header_row(raw_rows)
    header_row = raw_rows[header_idx]
    labels = _unique_labels(header_row)
    mapping = _map_columns(header_row)

    data_rows = raw_rows[header_idx + 1:]
    truncated = len([r for r in data_rows if any(_cell_text(c) for c in r)]) > MAX_ROWS

    # Tipo predominante de cada coluna (só para a UI alinhar e formatar)
    kinds: List[Dict[str, int]] = [{} for _ in labels]

    rows: List[Dict[str, Any]] = []
    seen: Dict[str, int] = {}
    counts = {"ok": 0, "duplicate_file": 0, "duplicate_db": 0, "invalid": 0}

    for offset, raw in enumerate(data_rows):
        if len(rows) >= MAX_ROWS:
            break
        row_number = header_idx + 2 + offset

        cells: Dict[str, Any] = {}
        mapped: Dict[str, str] = {}
        for idx, label in enumerate(labels):
            value = _cell_value(raw[idx]) if idx < len(raw) else None
            if value is None:
                continue
            cells[label] = value
            kind = _kind_of(raw[idx])
            kinds[idx][kind] = kinds[idx].get(kind, 0) + 1
            field = mapping.get(idx)
            if field and field not in mapped:
                mapped[field] = str(value).strip()

        if not cells:
            continue  # linha em branco: nada a preservar

        lead: Dict[str, Any] = {}
        for field in TEXT_FIELDS:
            text = mapped.get(field, "")
            if field == "linkedin_url":
                text = _clean_linkedin(text)
            if text:
                lead[field] = text[:_MAX_LEN.get(field, 255)]
        employees = parse_employee_count(mapped.get("employee_count", ""))
        if employees:
            lead["employee_count"] = employees
        domain, domain_source = _extract_domain(mapped)
        if domain:
            lead["domain"] = domain
            lead["website"] = f"https://{domain}"
            lead["domain_source"] = domain_source

        keys = identity_keys(lead)
        if not keys:
            counts["invalid"] += 1
            rows.append({
                "row_number": row_number, "status": "invalid",
                "reason": "Linha sem empresa, domínio ou LinkedIn para identificar.",
                "cells": cells, "lead": lead,
            })
            continue

        status, reason = "ok", None
        earlier = next((seen[k] for k in keys if k in seen), None)
        if earlier is not None:
            status = "duplicate_file"
            reason = f"Mesma empresa da linha {earlier} — será importada assim mesmo."
            counts["duplicate_file"] += 1
        else:
            counts["ok"] += 1
        for key in keys:
            seen.setdefault(key, row_number)

        rows.append({
            "row_number": row_number, "status": status, "reason": reason,
            "cells": cells, "lead": lead,
        })

    if not rows:
        raise SpreadsheetError(
            f'A aba "{sheet_name}" não tem nenhuma linha de dados abaixo do cabeçalho.'
        )

    columns = [
        {
            "label": label,
            "field": mapping.get(idx),
            "kind": max(kinds[idx], key=kinds[idx].get) if kinds[idx] else "text",
        }
        for idx, label in enumerate(labels)
    ]

    return {
        "sheet": sheet_name,
        "sheets": summaries,
        "columns": columns,
        "rows": rows,
        "total_rows": len(rows),
        "truncated": truncated,
        "counts": counts,
    }


# ── modelo de planilha para download ─────────────────────────────────────────

def build_template_xlsx() -> bytes:
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas"
    for col, (label, _) in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, sample in enumerate(TEMPLATE_SAMPLE, start=2):
        for col, value in enumerate(sample, start=1):
            ws.cell(row=row_idx, column=col, value=value)
    for col, width in enumerate([24, 26, 22, 22, 28, 20, 38, 16, 40], start=1):
        ws.column_dimensions[chr(64 + col)].width = width
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, dialect="excel")
    writer.writerow([label for label, _ in TEMPLATE_HEADERS])
    for sample in TEMPLATE_SAMPLE:
        writer.writerow(sample)
    return buf.getvalue().encode("utf-8-sig")
