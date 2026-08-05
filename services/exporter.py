import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List

HEADERS = [
    ("Empresa", "company_name"),
    ("Site", "website"),
    ("LinkedIn", "linkedin_url"),
    ("Setor", "sector"),
    ("Funcionários", "employee_count"),
    ("Localização", "location"),
    ("Descrição", "description"),
    ("Decisor", "decision_maker_name"),
    ("Cargo", "decision_maker_title"),
    ("LinkedIn Decisor", "decision_maker_linkedin"),
    ("Email Corporativo", "corporate_email"),
    ("Telefone", "phone"),
    ("Status", "status"),
    ("Data", "created_at"),
]

PRIMARY = "6C63FF"
HEADER_FONT_COLOR = "FFFFFF"
ALT_ROW = "F4F3FF"

# Caracteres que fazem Excel/Sheets tratarem o texto como fórmula.
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _sanitize(value: str) -> str:
    """
    Neutraliza injeção de fórmula na planilha.

    Nome de empresa, descrição e cargo vêm de sites de terceiros. Um valor
    começando com "=" é executado como fórmula ao abrir o arquivo — o campo
    coletado de um site qualquer viraria comando rodando na máquina de quem
    abre o export. O apóstrofo à frente força o Excel a tratar como texto.
    """
    if value and value[0] in _FORMULA_PREFIXES:
        return "'" + value
    return value


def _format_employee_count(value) -> str:
    """Converte o JSON consolidado {raw,min,max,band,exact} em texto legível."""
    if not isinstance(value, dict):
        return str(value) if value else ""
    if value.get("exact"):
        return str(value["exact"])
    if value.get("min") and value.get("max"):
        return f"{value['min']}–{value['max']}"
    if value.get("min"):
        return f"{value['min']}+"
    return value.get("band") or value.get("raw") or ""


def _field_value(lead, field):
    """Resolve o valor de uma coluna, incluindo campos derivados do 1º decisor."""
    if field.startswith("decision_maker_"):
        dms = getattr(lead, "decision_makers", None) or []
        if not dms:
            return ""
        dm = dms[0]
        return {
            "decision_maker_name": dm.name,
            "decision_maker_title": dm.title_found or dm.title_searched,
            "decision_maker_linkedin": dm.linkedin_url,
        }.get(field) or ""
    if field == "employee_count":
        return _format_employee_count(getattr(lead, field, None))
    return getattr(lead, field, None)


def _thin_border():
    side = Side(style="thin", color="DDDDDD")
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(leads: List) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Enriquecidos"

    # Header row
    for col_idx, (label, _) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.fill = PatternFill("solid", fgColor=PRIMARY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()

    ws.row_dimensions[1].height = 30

    # Data rows
    for row_idx, lead in enumerate(leads, start=2):
        fill = PatternFill("solid", fgColor=ALT_ROW) if row_idx % 2 == 0 else None
        for col_idx, (_, field) in enumerate(HEADERS, start=1):
            value = _field_value(lead, field)
            if value is None:
                value = ""
            elif hasattr(value, "strftime"):
                value = value.strftime("%d/%m/%Y %H:%M")
            text = _sanitize(str(value)) if value or value == 0 else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=text)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = _thin_border()
            if fill:
                cell.fill = fill

    # Column widths
    col_widths = [25, 30, 35, 20, 15, 25, 50, 25, 20, 35, 30, 18, 12, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_csv(leads: List) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, dialect="excel")
    writer.writerow([label for label, _ in HEADERS])
    for lead in leads:
        row = []
        for _, field in HEADERS:
            value = _field_value(lead, field)
            if value is None:
                row.append("")
            elif hasattr(value, "strftime"):
                row.append(value.strftime("%d/%m/%Y %H:%M"))
            else:
                row.append(_sanitize(str(value)))
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM para Excel abrir corretamente
