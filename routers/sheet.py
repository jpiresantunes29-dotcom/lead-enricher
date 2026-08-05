"""
A planilha dentro do sistema.

Serve o grid que substitui o Excel: as colunas originais do arquivo (lidas de
Lead.cells, na ordem em que o usuário criou) mais as colunas do sistema, que
são o resultado do enriquecimento. Editar uma célula da planilha mexe em
Lead.cells; editar uma coluna do sistema mexe no campo do Lead. As duas
famílias nunca se sobrescrevem — é isso que mantém o arquivo original fiel
depois do enriquecimento.

Busca, ordenação e filtro rodam em Python sobre o lote carregado (milhares de
linhas resolvem em milissegundos) porque consultar JSON em SQL muda de sintaxe
entre SQLite e Postgres, e a portabilidade aqui vale mais que o microssegundo.
"""
import io
import logging
import re
from datetime import date, datetime
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from middleware.auth import get_current_user
from models.database import ImportBatch, Lead, get_db
from models.schemas import (
    CellUpdate, ImportBatchOut, SheetColumn, SheetResponse,
    SheetRowCreate, SheetRowOut, SheetRowsDelete,
)
from services._utils import normalize_domain
from services.importer import parse_employee_count

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api", tags=["sheet"])

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Colunas do sistema — o que o enriquecimento preenche. label é o que aparece
# no grid; field é o atributo do Lead.
SYSTEM_COLUMNS: List[Dict[str, str]] = [
    {"label": "Domínio", "field": "domain", "kind": "text"},
    {"label": "Empresa (coletada)", "field": "company_name", "kind": "text"},
    {"label": "Site", "field": "website", "kind": "text"},
    {"label": "LinkedIn (coletado)", "field": "linkedin_url", "kind": "text"},
    {"label": "Setor", "field": "sector", "kind": "text"},
    {"label": "Localização", "field": "location", "kind": "text"},
    {"label": "Funcionários", "field": "employee_count", "kind": "text"},
    {"label": "Email corporativo", "field": "corporate_email", "kind": "text"},
    {"label": "Telefone", "field": "phone", "kind": "text"},
    {"label": "Provedor de e-mail", "field": "mx_provider", "kind": "text"},
    {"label": "Hospedagem", "field": "hosting_provider", "kind": "text"},
    {"label": "Score", "field": "score", "kind": "number"},
    {"label": "Prioridade", "field": "priority", "kind": "text"},
    {"label": "Estágio", "field": "stage", "kind": "text"},
    {"label": "Situação", "field": "status", "kind": "text"},
]

SYSTEM_BY_LABEL = {c["label"]: c for c in SYSTEM_COLUMNS}
# Campos editáveis à mão no grid (score/prioridade/situação são calculados)
EDITABLE_FIELDS = {
    "domain", "company_name", "website", "linkedin_url", "sector", "location",
    "employee_count", "corporate_email", "phone", "mx_provider", "stage",
}

STATUS_LABELS = {
    "imported": "Importado",
    "enriched": "Enriquecido",
    "partial": "Parcial",
    "failed": "Falhou",
}


def _employee_text(value: Any) -> str:
    """{raw,min,max,exact} → texto curto para a célula."""
    if not isinstance(value, dict):
        return str(value) if value not in (None, "") else ""
    if value.get("exact"):
        return str(value["exact"])
    if value.get("min") and value.get("max"):
        return f"{value['min']}–{value['max']}"
    if value.get("min"):
        return f"{value['min']}+"
    return value.get("band") or value.get("raw") or ""


def _system_value(lead: Lead, field: str) -> Any:
    value = getattr(lead, field, None)
    if field == "employee_count":
        return _employee_text(value)
    if field == "status":
        return STATUS_LABELS.get(value or "", value or "")
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _row_values(lead: Lead, columns: List[Dict]) -> Dict[str, Any]:
    cells = lead.cells or {}
    values: Dict[str, Any] = {col["label"]: cells.get(col["label"]) for col in columns}
    for col in SYSTEM_COLUMNS:
        values[col["label"]] = _system_value(lead, col["field"])
    return values


def _batch_columns(batch: Optional[ImportBatch]) -> List[Dict]:
    return list(batch.columns or []) if batch else []


def _batch_out(db: Session, user_id: str, batch: ImportBatch) -> ImportBatchOut:
    total = db.query(Lead).filter(
        Lead.user_id == user_id, Lead.import_batch_id == batch.id).count()
    enriched = db.query(Lead).filter(
        Lead.user_id == user_id, Lead.import_batch_id == batch.id,
        Lead.status.notin_(("imported", "failed"))).count()
    return ImportBatchOut(
        id=batch.id, filename=batch.filename, sheet_name=batch.sheet_name,
        row_count=total, enriched_count=enriched, created_at=batch.created_at,
    )


def _matches(values: Dict[str, Any], needle: str) -> bool:
    needle = needle.lower()
    return any(needle in str(v).lower() for v in values.values() if v not in (None, ""))


_LEADING_NUMBER = re.compile(r"^[^\d]{0,3}?(\d[\d.,]*)")


def _sort_key(value: Any):
    """
    Número ordena como número, texto como texto — e texto que começa com
    número ordena pelo número. Sem isso, uma coluna de funcionários com
    '39' e '51 a 200' misturados sairia em ordem alfabética.
    """
    if isinstance(value, bool):
        return (1, 0.0, str(value).lower())
    if isinstance(value, (int, float)):
        return (0, float(value), "")
    text = str(value).strip()
    match = _LEADING_NUMBER.match(text)
    if match:
        try:
            number = float(match.group(1).replace(".", "").replace(",", "."))
            return (0, number, text.lower())
        except ValueError:
            pass
    return (1, 0.0, text.lower())


def _sorted_rows(rendered: List, column: str, descending: bool) -> List:
    """
    Ordena mantendo as células vazias no fim nos dois sentidos — inverter a
    lista jogaria 700 linhas em branco para o topo, que é o oposto do que
    alguém quer ao clicar em "maior primeiro".
    """
    filled = [item for item in rendered if item[1].get(column) not in (None, "")]
    empty = [item for item in rendered if item[1].get(column) in (None, "")]
    filled.sort(key=lambda item: _sort_key(item[1].get(column)), reverse=descending)
    return filled + empty


@router.get("/sheet", response_model=SheetResponse)
def get_sheet(
    batch_id: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(100, ge=1, le=500),
    q: Optional[str] = Query(None),
    sort: Optional[str] = Query(None),
    dir: str = Query("asc", pattern="^(asc|desc)$"),
    only: Optional[str] = Query(None, pattern="^(pending|enriched|failed)$"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Uma página do grid, já com busca, filtro e ordenação aplicados."""
    user_id = current_user.get("sub")

    batches = (
        db.query(ImportBatch)
        .filter(ImportBatch.user_id == user_id, ImportBatch.status == "committed")
        .order_by(ImportBatch.created_at.desc())
        .all()
    )
    batch = None
    if batch_id and batch_id != "all":
        batch = next((b for b in batches if b.id == batch_id), None)
        if not batch:
            raise HTTPException(status_code=404, detail="Importação não encontrada.")
    elif batches and batch_id != "all":
        batch = batches[0]

    query = db.query(Lead).filter(Lead.user_id == user_id)
    if batch:
        query = query.filter(Lead.import_batch_id == batch.id)
    if only == "pending":
        query = query.filter(Lead.status == "imported")
    elif only == "enriched":
        query = query.filter(Lead.status.notin_(("imported", "failed")))
    elif only == "failed":
        query = query.filter(Lead.status == "failed")

    leads = query.order_by(Lead.sheet_row.asc(), Lead.id.asc()).all()
    columns = _batch_columns(batch)

    rendered = [(lead, _row_values(lead, columns)) for lead in leads]
    if q:
        rendered = [item for item in rendered if _matches(item[1], q)]

    if sort:
        rendered = _sorted_rows(rendered, sort, dir == "desc")

    total = len(rendered)
    start = (page - 1) * per_page
    window = rendered[start:start + per_page]

    return SheetResponse(
        batch=_batch_out(db, user_id, batch) if batch else None,
        batches=[_batch_out(db, user_id, b) for b in batches],
        columns=[SheetColumn(**col) for col in columns],
        system_columns=[SheetColumn(**col) for col in SYSTEM_COLUMNS],
        rows=[
            SheetRowOut(id=lead.id, sheet_row=lead.sheet_row, status=lead.status, values=values)
            for lead, values in window
        ],
        total=total,
        page=page,
        per_page=per_page,
        enrichable=sum(1 for lead, _ in rendered if lead.status == "imported"),
        pending_ids=[lead.id for lead, _ in rendered if lead.status == "imported"],
    )


@router.patch("/sheet/cell")
def update_cell(
    body: CellUpdate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Edita uma célula. Coluna do sistema vai para o campo do Lead; coluna da
    planilha vai para Lead.cells, preservando o arquivo original.
    """
    user_id = current_user.get("sub")
    lead = db.query(Lead).filter(Lead.id == body.lead_id, Lead.user_id == user_id).first()
    if not lead:
        raise HTTPException(status_code=404, detail="Linha não encontrada.")

    value = body.value
    if isinstance(value, str):
        value = value.strip()
    column = SYSTEM_BY_LABEL.get(body.column)

    if column:
        field = column["field"]
        if field not in EDITABLE_FIELDS:
            raise HTTPException(
                status_code=422,
                detail=f'A coluna "{body.column}" é calculada pelo sistema e não é editável.',
            )
        if field == "domain":
            value = normalize_domain(str(value or ""))
            lead.domain = value or None
            if value:
                lead.website = f"https://{value}"
                if not lead.raw_input_domain:
                    lead.raw_input_domain = value
        elif field == "employee_count":
            lead.employee_count = parse_employee_count(value)
        else:
            setattr(lead, field, value or None)
    else:
        cells = dict(lead.cells or {})
        if value in (None, ""):
            cells.pop(body.column, None)
        else:
            cells[body.column] = value
        lead.cells = cells
        flag_modified(lead, "cells")

    db.commit()
    db.refresh(lead)
    columns = []
    if lead.import_batch_id:
        batch = db.query(ImportBatch).filter(ImportBatch.id == lead.import_batch_id).first()
        columns = _batch_columns(batch)
    return {"success": True, "row": SheetRowOut(
        id=lead.id, sheet_row=lead.sheet_row, status=lead.status,
        values=_row_values(lead, columns),
    )}


@router.post("/sheet/row")
def create_row(
    body: SheetRowCreate,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Nova linha na planilha — o equivalente a digitar no fim do arquivo."""
    user_id = current_user.get("sub")
    batch = None
    if body.batch_id:
        batch = (
            db.query(ImportBatch)
            .filter(ImportBatch.id == body.batch_id, ImportBatch.user_id == user_id)
            .first()
        )
        if not batch:
            raise HTTPException(status_code=404, detail="Importação não encontrada.")

    columns = _batch_columns(batch)
    cells: Dict[str, Any] = {}
    lead_fields: Dict[str, Any] = {}
    for label, value in (body.values or {}).items():
        if value in (None, ""):
            continue
        system = SYSTEM_BY_LABEL.get(label)
        if system and system["field"] in EDITABLE_FIELDS:
            lead_fields[system["field"]] = value
        else:
            cells[label] = value
    # Colunas da planilha que também alimentam campos do Lead
    for col in columns:
        if col.get("field") and cells.get(col["label"]) and not lead_fields.get(col["field"]):
            lead_fields[col["field"]] = cells[col["label"]]

    domain = normalize_domain(str(lead_fields.get("domain") or ""))
    if lead_fields.get("employee_count"):
        lead_fields["employee_count"] = parse_employee_count(lead_fields["employee_count"])

    last_row = (
        db.query(Lead.sheet_row)
        .filter(Lead.user_id == user_id,
                Lead.import_batch_id == (batch.id if batch else None))
        .order_by(Lead.sheet_row.desc())
        .first()
    )
    next_row = ((last_row[0] or 1) + 1) if last_row else 2

    lead = Lead(
        user_id=user_id,
        raw_input_domain=domain or str(lead_fields.get("company_name") or ""),
        status="imported",
        stage="novo",
        cells=cells,
        import_batch_id=batch.id if batch else None,
        sheet_row=next_row,
        **{k: v for k, v in lead_fields.items() if k != "domain"},
    )
    if domain:
        lead.domain = domain
        lead.website = f"https://{domain}"
    db.add(lead)
    db.commit()
    db.refresh(lead)
    return {"success": True, "row": SheetRowOut(
        id=lead.id, sheet_row=lead.sheet_row, status=lead.status,
        values=_row_values(lead, columns),
    )}


@router.post("/sheet/rows/delete")
def delete_rows(
    body: SheetRowsDelete,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Remove linhas selecionadas (o cascade leva decisores e atividades)."""
    user_id = current_user.get("sub")
    if not body.ids:
        raise HTTPException(status_code=422, detail="Nenhuma linha selecionada.")
    leads = (
        db.query(Lead)
        .filter(Lead.user_id == user_id, Lead.id.in_(body.ids[:1000]))
        .all()
    )
    for lead in leads:
        db.delete(lead)
    db.commit()
    return {"success": True, "deleted": len(leads)}


@router.get("/sheet/export")
def export_sheet(
    batch_id: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    only: Optional[str] = Query(None, pattern="^(pending|enriched|failed)$"),
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """
    Exporta a planilha como está na tela: colunas originais na ordem original
    + colunas do sistema. É o caminho de volta para o Excel, sem perda.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    user_id = current_user.get("sub")
    batch = None
    if batch_id != "all":
        query_batch = db.query(ImportBatch).filter(
            ImportBatch.user_id == user_id, ImportBatch.status == "committed")
        if batch_id:
            batch = query_batch.filter(ImportBatch.id == batch_id).first()
            if not batch:
                raise HTTPException(status_code=404, detail="Importação não encontrada.")
        else:
            # Mesma escolha do grid: sem lote informado, exporta o mais recente
            batch = query_batch.order_by(ImportBatch.created_at.desc()).first()

    query = db.query(Lead).filter(Lead.user_id == user_id)
    if batch:
        query = query.filter(Lead.import_batch_id == batch.id)
    if only == "pending":
        query = query.filter(Lead.status == "imported")
    elif only == "enriched":
        query = query.filter(Lead.status.notin_(("imported", "failed")))
    elif only == "failed":
        query = query.filter(Lead.status == "failed")

    leads = query.order_by(Lead.sheet_row.asc(), Lead.id.asc()).all()
    columns = _batch_columns(batch)
    rendered = [(lead, _row_values(lead, columns)) for lead in leads]
    if q:
        rendered = [item for item in rendered if _matches(item[1], q)]
    if not rendered:
        raise HTTPException(status_code=404, detail="Nenhuma linha para exportar.")

    labels = [col["label"] for col in columns] + [col["label"] for col in SYSTEM_COLUMNS]

    wb = Workbook()
    ws = wb.active
    ws.title = (batch.sheet_name if batch else "Planilha")[:31] or "Planilha"

    for idx, label in enumerate(labels, start=1):
        cell = ws.cell(row=1, column=idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        # Colunas do sistema em outra cor: dá para ver o que é seu e o que é nosso
        is_system = idx > len(columns)
        cell.fill = PatternFill("solid", fgColor="0F766E" if is_system else "1D4ED8")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    for row_idx, (_, values) in enumerate(rendered, start=2):
        for col_idx, label in enumerate(labels, start=1):
            value = values.get(label)
            if isinstance(value, (dict, list)):
                value = str(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    for idx in range(1, len(labels) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 24
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    name = (batch.filename if batch else "planilha").rsplit(".", 1)[0][:60]
    return Response(
        content=buf.read(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{name}_leadenricher.xlsx"'},
    )
